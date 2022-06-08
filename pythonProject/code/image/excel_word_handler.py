# -*-coding:utf-8-*-
import datetime

from openpyxl import Workbook


def create_excel():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 42
    ws.append([1, 2, 3])
    ws['A2'] = datetime.datetime.now()
    wb.save('../../files/sample.xlsx')


def create_doc():
    pass


def main():
    create_excel()


if __name__ == '__main__':
    main()
